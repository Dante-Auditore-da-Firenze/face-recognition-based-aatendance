# import openpyxl module 
import openpyxl
from openpyxl import load_workbook

import pandas as pd
from datetime import datetime
import pytz

date1 = datetime.date(datetime.now())

ist = pytz.timezone('Asia/Calcutta')

time1 = datetime.time(datetime.now(ist))



writer = load_workbook("demo.xlsx")


sheet = writer.active 

c1 = sheet.cell(row = sheet.max_row + 2, column = 1) 

# writing values to cells 
c1.value = date1

c2 = sheet.cell(row = sheet.max_row, column = 3) 

# writing values to cells 
c2.value = time1

writer.save("demo.xlsx") 

def functionRecognize():
    import openpyxl
    from openpyxl import load_workbook
    import pandas as pd
    from datetime import datetime
    import pytz
    import cv2
    import numpy as np
    import sqlite3
    from keras.models import load_model
    from keras.preprocessing import image
    import numpy as np
    import os
    from os import listdir
    from os.path import isfile, join
    import re
    
    person={}
    n_id = os.listdir("./dataset/train")
    #print(n_id)
    for i,id1 in enumerate(n_id):
        person[i]=id1
    
    faceDetect=cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
    cam=cv2.VideoCapture(0)
    
    rec = load_model("./classifier.h5")
    font = cv2.FONT_HERSHEY_SIMPLEX
    

    def putAttendance(name):
        date1 = datetime.date(datetime.now())
        ist = pytz.timezone('Asia/Calcutta')
        time1 = datetime.time(datetime.now(ist))
        writer = load_workbook("Attendance.xlsx")
        sheet = writer.active 
        c1 = sheet.cell(row = sheet.max_row + 2, column = 1)
        c1.value = date1
        c2 = sheet.cell(row = sheet.max_row, column = 3) 
        c2.value = time1
        c3 = sheet.cell(row = sheet.max_row, column = 5)
        c3.value = name
        c4 = sheet.cell(row = sheet.max_row, column = 7)
        c4.value = "Present"
        writer.save("Attendance.xlsx")


    def getProfile(id):
        conn=sqlite3.connect("FaceBase.db")
    #    cmd="SELECT * FROM People WHERE ID="+str(id)
    #    cursor=conn.execute(cmd)
    #    profile=None
    #    for row in cursor:
    #        profile=row
    #    conn.close()
    #    return profile
        c = conn.cursor()
        c.execute("SELECT * FROM People WHERE Id=:id", {'id': id})
        ans = c.fetchone()
        c.close()
        return ans
    att_count = 0
    while(True):
        ret,img=cam.read();
        gray=cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
        faces=faceDetect.detectMultiScale(gray,1.3,5);
        for(x,y,w,h) in faces:
            cv2.rectangle(img,(x,y),(x+w,y+h),(0,255,0),2)
            input_im = img[y:y + h, x:x + w]
            input_im = cv2.resize(input_im, (100,100))
            input_im = input_im / 255.
            input_im = input_im.reshape(1,100,100,3)     
    # Get Prediction
            asd=rec.predict(input_im, 1, verbose = 0)
            maxp_index = np.argmax(asd)
            id = person[maxp_index]
            #print(asd[0][np.argmax(rec.predict(input_im, 1, verbose = 0))])
            profile=getProfile(id)
            conf=asd[0][maxp_index]
            #print(conf)
            if(att_count == 0 and conf*100 > 70):
                att_count+=1
                putAttendance(str(profile[1]))
            if(profile!=None):
                if conf*100 < 30:
                    cv2.putText(img, "Unknown",  (x,y-40), font, 1, (255,255,255), 3 )
                else:
                    cv2.putText(img, str(profile[1]) + " - " + str(profile[3]) + " ( {0:.2f}% )".format(round(100*conf, 2)) , (x,y-40), font, 1, (255,255,255), 3)
    #            cv2.cv.PutText(cv2.cv.fromarray(img),"Name : "+str(profile[1]),(x,y+h+20),font,(0,255,0));
    #            cv2.cv.PutText(cv2.cv.fromarray(img),"Age : "+str(profile[2]),(x,y+h+45),font,(0,255,0));
    #            cv2.cv.PutText(cv2.cv.fromarray(img),"Gender : "+str(profile[3]),(x,y+h+70),font,(0,255,0));
    #            cv2.cv.PutText(cv2.cv.fromarray(img),"Criminal Records : "+str(profile[4]),(x,y+h+95),font,(0,0,255));
        cv2.imshow("Face",img);
        if(cv2.waitKey(1)==ord('q')):
            break;
    cam.release()
    cv2.destroyAllWindows()